import csv
import datetime
import io
from pathlib import Path

from django import forms
from django.conf import settings
from django.core.exceptions import ValidationError

from ..layout.models import FeatureArticle, Homepage
from ..newsletter.models import Newsletter
from ..submissions.models import Article, Author, Issue, Journal, Review
from .models import InboundEmail, PlankaBoardBackgroundAsset, SubscriberCSV


def csv_size(file):
    limit = 1 * 1024 * 1024
    if file.size > limit:
        raise ValidationError({"file": "File too large. Size should not exceed 1 megabyte."})


DELIMITERS = [",", ";", "\t", " "]
EMAIL_HEADER_CANDIDATES = {"email", "email address", "e-mail", "e-mail address", "mail"}


class PreviewTable:
    def __init__(self, fieldnames, rows):
        self.fieldnames = fieldnames
        self._rows = rows

    def __iter__(self):
        return iter(self._rows)


def _normalize_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def _looks_like_email_header(value):
    normalized = (value or "").strip().lower()
    return normalized in EMAIL_HEADER_CANDIDATES


def _build_preview_from_rows(rows, user_header=None):
    if not rows:
        return {"preview": PreviewTable(["Column 1"], []), "has_header": False}

    first_row = rows[0]
    guessed_header = any(_looks_like_email_header(value) for value in first_row) and all(
        "@" not in (value or "") for value in first_row
    )

    has_header = guessed_header if user_header is None else bool(user_header)

    max_cols = max(len(row) for row in rows) if rows else 1
    padded_rows = [row + [""] * (max_cols - len(row)) for row in rows]

    if has_header:
        fieldnames = [value or f"Column {idx + 1}" for idx, value in enumerate(padded_rows[0])]
        data_rows = padded_rows[1:]
    else:
        fieldnames = [f"Column {idx + 1}" for idx in range(max_cols)]
        data_rows = padded_rows

    dict_rows = [dict(zip(fieldnames, row, strict=False)) for row in data_rows]
    return {"preview": PreviewTable(fieldnames, dict_rows), "has_header": has_header}


def _peek_xlsx(file, user_header=None):
    try:
        from openpyxl import load_workbook
    except Exception as error:
        raise ValidationError({"file": f"XLSX support requires openpyxl: {error}"})

    try:
        file.seek(0)
        workbook = load_workbook(filename=file, read_only=True, data_only=True)
    except Exception as error:
        raise ValidationError({"file": f"Not a valid XLSX file: {error}"})

    worksheet = workbook.active
    rows = []
    for raw_row in worksheet.iter_rows(min_row=1, max_row=25, values_only=True):
        row = [_normalize_cell(value) for value in raw_row]
        while row and row[-1] == "":
            row.pop()
        if not row:
            continue
        rows.append(row)

    return _build_preview_from_rows(rows, user_header=user_header)


def peek_csv(file, user_header=None):
    filename = Path(getattr(file, "name", "")).suffix.lower()
    if filename == ".xlsx":
        return _peek_xlsx(file, user_header=user_header)

    try:
        file.seek(0)
        decoded_file = file.read(1024).decode("UTF-8-SIG")
    except UnicodeDecodeError as error:
        print(f"Error handling uploaded CSV: {error}")
        raise ValidationError({"file": "Not a valid CSV file"})

    try:
        dialect = csv.Sniffer().sniff(decoded_file, DELIMITERS)
    except csv.Error as error:
        for delimiter in DELIMITERS:
            if delimiter in decoded_file:
                print(f"Error handling uploaded CSV: {error}")
                raise ValidationError({"file": "Not a valid CSV file"})
        # No delimiter found; likely single-column file
        dialect = csv.excel

    has_header = csv.Sniffer().has_header(decoded_file)

    # Determine column number and names
    delimiter = str(dialect.delimiter)
    fieldnames = decoded_file.split("\n")[0].split(delimiter)

    # If user has selected header
    if user_header is not None:
        has_header = user_header

    if not has_header:
        column_count = len(fieldnames)
        fieldnames = []
        for i in range(column_count):
            fieldnames.append(f"Column {i+1}")
    else:
        fieldnames = None  # Allow DictReader to use the first row as fieldnames

    io_string = io.StringIO(decoded_file)
    preview_rows = list(csv.DictReader(io_string, fieldnames=fieldnames, dialect=dialect))
    file.seek(0)

    return {
        "preview": PreviewTable(preview_rows[0].keys() if preview_rows else fieldnames or [], preview_rows),
        "has_header": has_header,
    }


class SubscriberCSVForm(forms.ModelForm):
    class Meta:
        model = SubscriberCSV
        fields = [
            "name",
            "file",
        ]

    def clean(self):
        cleaned_data = super().clean()

        # File is already opened by Django
        file = cleaned_data["file"]

        # Validate and preview the CSV
        csv_size(file)
        csv_preview = peek_csv(file)

        cleaned_data.update(csv_preview)
        return cleaned_data


class HeaderForm(forms.Form):
    header = forms.BooleanField(label="The first row of this CSV is a column heading", required=False)


class NewsletterTestSendForm(forms.Form):
    email = forms.EmailField(label="Send test email to")

    def clean_email(self):
        return self.cleaned_data["email"].lower().strip()


class NewsletterCreateForm(forms.ModelForm):
    class Meta:
        model = Newsletter
        fields = [
            "subject",
            "issue",
            "content_heading",
            "content",
            "header_image",
            "non_featured_review_count",
            "ready_to_send",
        ]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        issue_qs = Issue.objects.exclude(active=False).order_by("-date", "-pk")
        self.fields["issue"].queryset = issue_qs
        self.fields["issue"].required = False

        if not self.instance.pk and issue_qs.exists():
            self.fields["issue"].initial = issue_qs.first()

    def clean_issue(self):
        issue = self.cleaned_data.get("issue")
        if issue:
            return issue

        fallback_issue = Issue.objects.exclude(active=False).order_by("-date", "-pk").first()
        if fallback_issue:
            return fallback_issue

        raise ValidationError("Please select an issue.")


class InboundAnymailEmailForm(forms.ModelForm):
    class Meta:
        model = InboundEmail
        fields = [
            "sender",
            "recipient",
            "header_sender",
            "header_recipients",
            "subject",
            "body",
            "body_html",
            "sent_timestamp",
            "attachments",
            "email_file",
        ]

    def clean_attachments(self):
        return bool(self.cleaned_data["attachments"])


# Data entry
# ----------------------------------------------------


class FeatureArticleForm(forms.ModelForm):
    class Meta:
        model = FeatureArticle
        fields = [
            "title",
            "body",
            "image",
        ]


class HomepageForm(forms.ModelForm):
    # Create a new Homepage
    # Default HomepagePage with its associated FeatureArticle
    # Method for creating a new FeatureArticle with the following:
    # - Title
    # - Body
    # - Image
    class Meta:
        model = Homepage
        fields = [
            "issue",
            "override_main",
        ]


class ArticleForm(forms.ModelForm):
    class Meta:
        model = Article
        fields = ["name", "journal", "year", "citation", "url", "tags_string"]


class ReviewForm(forms.ModelForm):
    class Meta:
        model = Review
        fields = ["author", "body", "publish_date", "is_featured", "feature_image"]


class IssueForm(forms.ModelForm):
    class Meta:
        model = Issue
        fields = ["name", "date", "body"]


class IssueBuilderIssueForm(forms.ModelForm):
    date = forms.DateField(
        required=False,
        input_formats=["%Y-%m", "%Y-%m-%d"],
        widget=forms.DateInput(attrs={"type": "month"}),
        help_text="Select issue month and year.",
    )

    class Meta:
        model = Issue
        fields = ["name", "date", "body"]

    def clean_date(self):
        value = self.cleaned_data.get("date")
        if value:
            return value.replace(day=1)
        return value


class IssueBuilderReviewForm(forms.Form):
    AUTHOR_MODE_EXISTING = "existing"
    AUTHOR_MODE_NEW = "new"
    AUTHOR_MODE_CHOICES = [
        (AUTHOR_MODE_EXISTING, "Select existing author"),
        (AUTHOR_MODE_NEW, "Create new author"),
    ]

    ARTICLE_MODE_NEW = "new"
    ARTICLE_MODE_EXISTING = "existing"
    ARTICLE_MODE_CHOICES = [
        (ARTICLE_MODE_NEW, "Create new article"),
        (ARTICLE_MODE_EXISTING, "Select existing article"),
    ]

    article_mode = forms.ChoiceField(choices=ARTICLE_MODE_CHOICES, initial=ARTICLE_MODE_NEW)
    existing_article = forms.ModelChoiceField(
        queryset=Article.objects.order_by("name"),
        required=False,
        help_text="Use an existing article instead of creating a new one.",
    )

    article_name = forms.CharField(required=False)
    article_journal = forms.ModelChoiceField(queryset=Journal.objects.order_by("name"), required=False)
    article_year = forms.IntegerField(required=False, min_value=1900)
    article_citation = forms.CharField(required=False, widget=forms.Textarea(attrs={"rows": 2}))
    article_url = forms.URLField(required=False)
    article_tags_string = forms.CharField(required=False)

    author_mode = forms.ChoiceField(choices=AUTHOR_MODE_CHOICES, initial=AUTHOR_MODE_EXISTING, required=False)
    author = forms.ModelChoiceField(queryset=Author.objects.order_by("name"), required=False)
    new_author_title = forms.CharField(required=False, initial="Dr")
    new_author_name = forms.CharField(required=False)
    body = forms.CharField(widget=forms.Textarea(attrs={"rows": 6}), required=True)
    publish_date = forms.DateField(required=False, widget=forms.DateInput(attrs={"type": "date"}))
    is_featured = forms.BooleanField(required=False)
    feature_image = forms.ImageField(required=False)

    def __init__(self, *args, issue=None, review=None, **kwargs):
        self.issue = issue
        self.review = review
        super().__init__(*args, **kwargs)

        if review:
            self.fields["article_mode"].initial = self.ARTICLE_MODE_EXISTING
            self.fields["existing_article"].initial = review.article
            self.fields["author_mode"].initial = self.AUTHOR_MODE_EXISTING
            self.fields["author"].initial = review.author
            self.fields["body"].initial = review.body
            self.fields["publish_date"].initial = review.publish_date
            self.fields["is_featured"].initial = review.is_featured

    @property
    def max_featured_reviews(self):
        return int(getattr(settings, "ISSUE_BUILDER_MAX_FEATURED_REVIEWS", 2))

    def clean(self):
        cleaned_data = super().clean()
        article_mode = cleaned_data.get("article_mode")

        if article_mode == self.ARTICLE_MODE_EXISTING:
            if not cleaned_data.get("existing_article"):
                self.add_error("existing_article", "Select an existing article.")
        else:
            if not cleaned_data.get("article_name"):
                self.add_error("article_name", "Article title is required.")

        author_mode = cleaned_data.get("author_mode") or self.AUTHOR_MODE_EXISTING
        cleaned_data["author_mode"] = author_mode
        if author_mode == self.AUTHOR_MODE_NEW:
            if not cleaned_data.get("new_author_name"):
                self.add_error("new_author_name", "Author name is required.")
        else:
            if not cleaned_data.get("author"):
                self.add_error("author", "Select an author.")

        is_featured = cleaned_data.get("is_featured")
        if is_featured and self.issue:
            featured_qs = self.issue.reviews.filter(is_featured=True)
            if self.review:
                featured_qs = featured_qs.exclude(pk=self.review.pk)
            if featured_qs.count() >= self.max_featured_reviews:
                self.add_error(
                    "is_featured",
                    f"Only {self.max_featured_reviews} featured reviews are allowed per issue.",
                )

        return cleaned_data

    def save(self):
        if not self.issue:
            raise ValidationError("Issue is required before adding a review.")

        article_mode = self.cleaned_data["article_mode"]

        if article_mode == self.ARTICLE_MODE_EXISTING:
            article = self.cleaned_data["existing_article"]
        else:
            default_year = datetime.date.today().year
            article = Article.objects.create(
                name=self.cleaned_data["article_name"],
                journal=self.cleaned_data.get("article_journal"),
                year=self.cleaned_data.get("article_year") or default_year,
                citation=self.cleaned_data.get("article_citation") or "",
                url=self.cleaned_data.get("article_url"),
                tags_string=self.cleaned_data.get("article_tags_string") or "",
            )

        author_mode = self.cleaned_data["author_mode"]
        if author_mode == self.AUTHOR_MODE_NEW:
            author = Author.objects.create(
                title=self.cleaned_data.get("new_author_title") or "Dr",
                name=self.cleaned_data["new_author_name"],
            )
        else:
            author = self.cleaned_data["author"]

        review = self.review if self.review else Review()
        review.article = article
        review.author = author
        review.body = self.cleaned_data["body"]
        review.publish_date = self.cleaned_data.get("publish_date")
        review.is_featured = self.cleaned_data.get("is_featured", False)
        if self.cleaned_data.get("feature_image"):
            review.feature_image = self.cleaned_data["feature_image"]
        review.save()

        self.issue.reviews.add(review)
        return review


class PlankaProjectSetupForm(forms.Form):
    project_name = forms.CharField(max_length=128)
    background_asset = forms.ModelChoiceField(
        queryset=PlankaBoardBackgroundAsset.objects.none(),
        required=False,
        empty_label="No background image",
        help_text="Choose a previously uploaded background image.",
    )
    background_upload = forms.ImageField(
        required=False,
        help_text="Or upload a new image (will be converted to WebP, max 1920×1080).",
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["background_asset"].queryset = PlankaBoardBackgroundAsset.objects.order_by("name")

    def clean_project_name(self):
        return self.cleaned_data["project_name"].strip()

    def clean(self):
        cleaned_data = super().clean()
        background_asset = cleaned_data.get("background_asset")
        background_upload = cleaned_data.get("background_upload")

        if background_asset and background_upload:
            self.add_error("background_upload", "Choose either an existing image or upload a new one, not both.")

        return cleaned_data


class PlankaApiKeyForm(forms.Form):
    api_key = forms.CharField(
        required=True,
        widget=forms.PasswordInput(render_value=False),
        help_text="Create a key in Planka (Profile/Settings → API key) and paste it here.",
    )
    issue_id = forms.IntegerField(required=False, widget=forms.HiddenInput())

    def clean_api_key(self):
        value = (self.cleaned_data.get("api_key") or "").strip()
        if not value:
            raise ValidationError("API key is required.")
        return value


class PlankaProjectBackgroundForm(forms.Form):
    background_asset = forms.ModelChoiceField(
        queryset=PlankaBoardBackgroundAsset.objects.none(),
        required=False,
        empty_label="Keep current background",
        help_text="Choose a previously uploaded background image.",
    )
    background_upload = forms.ImageField(
        required=False,
        help_text="Or upload a new image (will be converted to WebP, max 1920×1080).",
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["background_asset"].queryset = PlankaBoardBackgroundAsset.objects.order_by("name")

    def clean(self):
        cleaned_data = super().clean()
        background_asset = cleaned_data.get("background_asset")
        background_upload = cleaned_data.get("background_upload")

        if background_asset and background_upload:
            self.add_error("background_upload", "Choose either an existing image or upload a new one, not both.")

        return cleaned_data


class PlankaProjectNameForm(forms.Form):
    project_name = forms.CharField(max_length=128)

    def clean_project_name(self):
        return (self.cleaned_data.get("project_name") or "").strip()
