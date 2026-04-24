import csv
import datetime
import io
import logging
from pathlib import Path

from django import forms
from django.apps import apps
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.utils import timezone

from config.celery_app import app as celery_app

logger = logging.getLogger(__name__)


class MultipleEmailColumnsException(Exception):
    def __init__(self, message):
        self.message = message
        super().__init__(self.message)


class NoEmailColumnsFoundException(Exception):
    def __init__(self, message):
        self.message = message
        super().__init__(self.message)


def get_celery_subscriber_form():
    # Wrapped in fuction to avoid circular import

    Subscriber = apps.get_model("newsletter", "Subscriber")

    class CelerySubscriberForm(forms.ModelForm):
        class Meta:
            model = Subscriber
            fields = [
                "email",
                "subscribed",
                "from_csv",
            ]

        def clean_email(self):
            email = self.cleaned_data["email"].lower()
            if Subscriber.objects.filter(email__iexact=email).exists():
                raise forms.ValidationError(f"Email {email} is already a subscriber")
            return email

    return CelerySubscriberForm


DELIMITERS = [",", ";", "\t", " "]
EMAIL_HEADER_CANDIDATES = {"email", "email address", "e-mail", "e-mail address", "mail"}


def _normalize_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def _iter_csv_rows(file_obj):
    decoded_file = file_obj.read().decode("UTF-8-SIG")
    try:
        dialect = csv.Sniffer().sniff(decoded_file, DELIMITERS)
    except csv.Error:
        for delimiter in DELIMITERS:
            if delimiter in decoded_file:
                raise ValidationError("Not a valid CSV file")
        dialect = csv.excel

    io_string = io.StringIO(decoded_file)
    for row in csv.reader(io_string, dialect=dialect):
        yield [_normalize_cell(value) for value in row]


def _iter_xlsx_rows(file_obj):
    try:
        from openpyxl import load_workbook
    except Exception as error:
        raise ValidationError(f"XLSX support requires openpyxl: {error}")

    workbook = load_workbook(filename=file_obj, read_only=True, data_only=True)
    worksheet = workbook.active
    for row in worksheet.iter_rows(values_only=True):
        yield [_normalize_cell(value) for value in row]


def _load_rows(subscriber_csv):
    suffix = Path(subscriber_csv.file.name).suffix.lower()
    with subscriber_csv.file.open("rb") as file_obj:
        if suffix == ".xlsx":
            return [row for row in _iter_xlsx_rows(file_obj) if any(value for value in row)]
        return [row for row in _iter_csv_rows(file_obj) if any(value for value in row)]


def _find_header_email_column_index(headers):
    matches = []
    for idx, raw_header in enumerate(headers):
        normalized = (raw_header or "").strip().lower()
        if normalized in EMAIL_HEADER_CANDIDATES:
            matches.append(idx)

    if len(matches) > 1:
        raise MultipleEmailColumnsException("Multiple possible email header columns detected.")
    return matches[0] if matches else None


def _find_data_email_column_index(rows):
    if not rows:
        return None

    sample_rows = rows[: min(50, len(rows))]
    max_cols = max(len(row) for row in sample_rows)
    valid_counts = [0] * max_cols

    for row in sample_rows:
        for idx in range(max_cols):
            value = row[idx].strip().lower() if idx < len(row) else ""
            if not value:
                continue
            try:
                validate_email(value)
                valid_counts[idx] += 1
            except ValidationError:
                continue

    best_score = max(valid_counts) if valid_counts else 0
    if best_score == 0:
        return None

    winners = [idx for idx, score in enumerate(valid_counts) if score == best_score]
    if len(winners) > 1:
        raise MultipleEmailColumnsException("Multiple columns contain valid emails.")

    return winners[0]


def process_subscriber_csv_record(subscriber_csv):
    from spanza_journal_watch.newsletter.models import Subscriber

    rows = _load_rows(subscriber_csv)
    if not rows:
        raise NoEmailColumnsFoundException("No rows found in uploaded file")

    has_header = bool(subscriber_csv.header)
    headers = rows[0] if has_header else []
    data_rows = rows[1:] if has_header else rows

    email_col_index = _find_header_email_column_index(headers) if has_header else None
    if email_col_index is None:
        email_col_index = _find_data_email_column_index(data_rows)

    if email_col_index is None:
        raise NoEmailColumnsFoundException("Could not detect an email column")

    rows_parsed = 0
    records_added = 0
    invalid_email_count = 0
    duplicate_in_file_count = 0
    already_subscribed_count = 0
    seen_in_file = set()

    for row in data_rows:
        rows_parsed += 1
        email = row[email_col_index].strip().lower() if email_col_index < len(row) else ""
        if not email:
            invalid_email_count += 1
            continue

        try:
            validate_email(email)
        except ValidationError:
            invalid_email_count += 1
            continue

        if email in seen_in_file:
            duplicate_in_file_count += 1
            continue
        seen_in_file.add(email)

        if Subscriber.objects.filter(email__iexact=email).exists():
            already_subscribed_count += 1
            continue

        Subscriber.objects.create(
            email=email,
            subscribed=True,
            from_csv=subscriber_csv,
            source=Subscriber.Source.CSV_IMPORT,
        )
        records_added += 1

    subscriber_csv.processed = True
    subscriber_csv.row_count = rows_parsed
    subscriber_csv.email_added_count = records_added
    subscriber_csv.save(update_fields=["processed", "row_count", "email_added_count", "modified"])

    return {
        "rows_parsed": rows_parsed,
        "records_added": records_added,
        "records_skipped": invalid_email_count + duplicate_in_file_count + already_subscribed_count,
        "invalid_email_count": invalid_email_count,
        "duplicate_in_file_count": duplicate_in_file_count,
        "already_subscribed_count": already_subscribed_count,
        "email_column": headers[email_col_index]
        if has_header and email_col_index < len(headers)
        else f"Column {email_col_index + 1}",
    }


@celery_app.task()
def process_subscriber_csv(subscriber_csv_pk):
    # Prevent circular import
    from .models import SubscriberCSV

    try:
        subscriber_csv = SubscriberCSV.objects.get(pk=subscriber_csv_pk)
    except SubscriberCSV.DoesNotExist:
        return None

    subscriber_csv.task_state = SubscriberCSV.TASK_STATE_RUNNING
    subscriber_csv.task_note = "Processing subscriber list..."
    subscriber_csv.save(update_fields=["task_state", "task_note", "modified"])

    try:
        summary = process_subscriber_csv_record(subscriber_csv)
    except Exception as error:
        logger.exception("Subscriber CSV processing failed (pk=%s)", subscriber_csv_pk)
        SubscriberCSV.objects.filter(pk=subscriber_csv_pk).update(
            task_state=SubscriberCSV.TASK_STATE_ERROR,
            task_note=f"Import failed: {error}",
            modified=timezone.now(),
        )
        raise

    SubscriberCSV.objects.filter(pk=subscriber_csv_pk).update(
        task_state=SubscriberCSV.TASK_STATE_SUCCESS,
        task_note=f"Added {summary.get('records_added', 0)} subscriber(s).",
        task_summary=summary,
        modified=timezone.now(),
    )
    return summary


@celery_app.task(bind=True)
def run_pubmed_batch_import_task(self, batch_id):
    from .models import FetchLog, PubmedImportBatch
    from .pubmed import PubmedAPIError
    from .pubmed_cache import populate_pubmed_batch_from_cache
    from .views import _safe_planka_error

    batch = PubmedImportBatch.objects.get(pk=batch_id)
    batch.task_state = PubmedImportBatch.TASK_STATE_RUNNING
    batch.task_id = self.request.id or batch.task_id
    batch.task_note = "Fetching articles from PubMed..."
    batch.save(update_fields=["task_state", "task_id", "task_note", "modified"])

    fetch_log = FetchLog.objects.create(
        task_type=FetchLog.TASK_BATCH_IMPORT,
        celery_task_id=self.request.id or "",
        details={"batch_id": batch_id},
    )

    watched_journals = list(batch.watched_journals.filter(active=True))
    if not watched_journals:
        batch.task_state = PubmedImportBatch.TASK_STATE_ERROR
        batch.task_note = "No active watched journals on this batch."
        batch.save(update_fields=["task_state", "task_note", "modified"])
        fetch_log.finish(FetchLog.STATUS_ERROR, error_message=batch.task_note)
        return {"status": "error", "note": batch.task_note}

    try:
        populate_pubmed_batch_from_cache(batch, watched_journals)
        batch.task_state = PubmedImportBatch.TASK_STATE_SUCCESS
        batch.task_note = f"Loaded {batch.result_count} cached article(s) into the intake batch."
        batch.save(update_fields=["task_state", "task_note", "modified"])
        fetch_log.finish(
            FetchLog.STATUS_SUCCESS,
            journal_count=len(watched_journals),
            articles_created=batch.result_count,
            details={"batch_id": batch_id, "result_count": batch.result_count},
        )
        return {"status": "success", "count": batch.result_count}
    except PubmedAPIError as error:
        batch.task_state = PubmedImportBatch.TASK_STATE_ERROR
        batch.task_note = f"PubMed fetch failed: {_safe_planka_error(error)}"
        batch.save(update_fields=["task_state", "task_note", "modified"])
        logger.error("PubMed import batch %s failed: %s", batch_id, error)
        fetch_log.finish(FetchLog.STATUS_ERROR, error_message=str(error))
        return {"status": "error", "note": batch.task_note}


@celery_app.task(bind=True)
def refresh_pubmed_journal_cache_task(self, from_month=None, to_month=None):
    from .models import FetchLog
    from .pubmed_cache import default_pubmed_cache_window, refresh_pubmed_journal_cache

    if from_month and to_month:
        from_month_value = datetime.date.fromisoformat(from_month)
        to_month_value = datetime.date.fromisoformat(to_month)
    else:
        from_month_value, to_month_value = default_pubmed_cache_window(timezone.now().date())

    fetch_log = FetchLog.objects.create(
        task_type=FetchLog.TASK_CACHE_REFRESH,
        celery_task_id=self.request.id or "",
        details={"from_month": from_month_value.isoformat(), "to_month": to_month_value.isoformat()},
    )

    try:
        stats = refresh_pubmed_journal_cache(from_month=from_month_value, to_month=to_month_value)
        logger.info(
            "Refreshed PubMed journal cache for %s to %s across %s journals",
            from_month_value,
            to_month_value,
            stats["journal_count"],
        )
        fetch_log.finish(
            FetchLog.STATUS_SUCCESS,
            journal_count=stats.get("journal_count", 0),
            articles_created=stats.get("created_links", 0),
            articles_touched=stats.get("touched_links", 0),
            details={
                "from_month": from_month_value.isoformat(),
                "to_month": to_month_value.isoformat(),
                **stats,
            },
        )
        return {
            "status": "success",
            "from_month": from_month_value.isoformat(),
            "to_month": to_month_value.isoformat(),
            **stats,
        }
    except Exception as exc:
        logger.error("PubMed cache refresh failed: %s", exc)
        fetch_log.finish(FetchLog.STATUS_ERROR, error_message=str(exc))
        raise


@celery_app.task(bind=True)
def refresh_mesh_terms_task(self):
    """Re-fetch PubMed metadata for articles missing MeSH terms, then auto-tag."""
    from .models import FetchLog, PubmedArticle
    from .pubmed_cache import build_pubmed_client, fill_missing_article_metadata

    fetch_log = FetchLog.objects.create(
        task_type=FetchLog.TASK_CACHE_REFRESH,
        celery_task_id=self.request.id or "",
        details={"type": "mesh_refresh"},
    )

    try:
        client = build_pubmed_client()
        candidates = []
        for article in PubmedArticle.objects.filter(pmid__isnull=False).iterator():
            if not (article.metadata_json or {}).get("mesh_terms"):
                candidates.append(article)

        pmid_to_article = {a.pmid: a for a in candidates}
        pmid_list = list(pmid_to_article.keys())
        batch_size = 200
        updated = 0
        still_empty = 0
        errors = 0

        for i in range(0, len(pmid_list), batch_size):
            batch = pmid_list[i : i + batch_size]
            try:
                payloads = client.fetch_articles(batch)
            except Exception:
                logger.exception("MeSH refresh: failed batch at offset %d", i)
                errors += len(batch)
                continue

            fetched = {}
            for payload in payloads:
                pmid = (payload.get("pmid") or "").strip()
                if pmid:
                    fetched[pmid] = payload

            for pmid in batch:
                article = pmid_to_article[pmid]
                payload = fetched.get(pmid)
                if not payload:
                    still_empty += 1
                    continue
                incoming_mesh = (payload.get("metadata_json") or {}).get("mesh_terms", [])
                if incoming_mesh:
                    fill_missing_article_metadata(article, payload)
                    updated += 1
                else:
                    still_empty += 1

        stats = {
            "type": "mesh_refresh",
            "candidates": len(candidates),
            "updated": updated,
            "still_empty": still_empty,
            "errors": errors,
        }
        fetch_log.finish(FetchLog.STATUS_SUCCESS, details=stats)
        logger.info("MeSH refresh complete: %s", stats)
        return stats
    except Exception as exc:
        logger.error("MeSH refresh failed: %s", exc)
        fetch_log.finish(FetchLog.STATUS_ERROR, error_message=str(exc))
        raise


@celery_app.task
def compute_tag_clusters_task():
    """Recompute tag co-occurrence clusters and cache the result."""
    from collections import defaultdict
    from itertools import combinations

    from django.core.cache import cache

    from spanza_journal_watch.submissions.management.commands.compute_tag_clusters import (
        CACHE_KEY,
        CACHE_TIMEOUT,
        SIMILARITY_THRESHOLD,
    )
    from spanza_journal_watch.submissions.models import Tag

    tag_articles = {}
    for tag in Tag.objects.filter(active=True, curated=True):
        article_ids = set(tag.articles.values_list("id", flat=True))
        if article_ids:
            tag_articles[tag.id] = article_ids

    adjacency = defaultdict(set)
    for (a_id, a_articles), (b_id, b_articles) in combinations(tag_articles.items(), 2):
        overlap = len(a_articles & b_articles)
        if overlap == 0:
            continue
        min_size = min(len(a_articles), len(b_articles))
        if overlap / min_size >= SIMILARITY_THRESHOLD:
            adjacency[a_id].add(b_id)
            adjacency[b_id].add(a_id)

    visited = set()
    clusters = []
    for tag_id in tag_articles:
        if tag_id in visited:
            continue
        component = set()
        queue = [tag_id]
        while queue:
            current = queue.pop()
            if current in visited:
                continue
            visited.add(current)
            component.add(current)
            queue.extend(adjacency.get(current, set()) - visited)
        if len(component) > 1:
            clusters.append(sorted(component))

    clusters.sort(key=len, reverse=True)
    cache.set(CACHE_KEY, clusters, CACHE_TIMEOUT)
    logger.info("Tag clusters recomputed: %d clusters from %d tags", len(clusters), len(tag_articles))
    return {"clusters": len(clusters), "tags": len(tag_articles)}


@celery_app.task(bind=True)
def run_pubmed_batch_push_task(self, batch_id, push_scope="selected"):
    from django.conf import settings
    from django.utils import timezone

    from .models import PubmedImportBatch
    from .planka import PlankaAPIError
    from .views import (
        _attach_journal_label_to_card,
        _build_planka_client,
        _build_pubmed_planka_card,
        _ensure_planka_board_mappings,
        _get_board_label_map,
        _get_board_list_type_map,
        _get_issue_planka_candidates_list,
        _is_planka_card_archived,
        _is_planka_card_not_found_error,
        _is_planka_list_not_found_error,
        _safe_planka_error,
    )

    batch = PubmedImportBatch.objects.get(pk=batch_id)
    batch.task_state = PubmedImportBatch.TASK_STATE_RUNNING
    batch.task_id = self.request.id or batch.task_id
    batch.task_note = "Pushing staged articles to Planka..."
    batch.save(update_fields=["task_state", "task_id", "task_note", "modified"])

    issue, binding, list_error = _get_issue_planka_candidates_list(batch, require_candidates_list=False)
    if list_error:
        batch.task_state = PubmedImportBatch.TASK_STATE_ERROR
        batch.task_note = list_error
        batch.save(update_fields=["task_state", "task_note", "modified"])
        return {"status": "error", "note": list_error}

    if push_scope == "filtered":
        target_rows = list(batch.batch_articles.select_related("article", "issue").filter(is_selected=True))
    else:
        target_rows = list(batch.batch_articles.select_related("article", "issue").filter(is_selected=True))

    if not target_rows:
        batch.task_state = PubmedImportBatch.TASK_STATE_SUCCESS
        batch.task_note = "No staged articles available to push."
        batch.save(update_fields=["task_state", "task_note", "modified"])
        return {"status": "success", "note": batch.task_note}

    try:
        client = _build_planka_client()
        _ensure_planka_board_mappings(client=client, binding=binding)
        label_cache = _get_board_label_map(client=client, board_id=binding.board_id)
        list_type_map = _get_board_list_type_map(client=client, board_id=binding.board_id)
    except PlankaAPIError as error:
        safe_error = _safe_planka_error(error)
        batch.task_state = PubmedImportBatch.TASK_STATE_ERROR
        if "board not found" in safe_error.lower():
            batch.task_note = "Linked Planka board was not found. Re-link this issue to a valid Planka project/board."
        else:
            batch.task_note = f"Could not prepare Planka board: {safe_error}"
        batch.save(update_fields=["task_state", "task_note", "modified"])
        logger.error("Planka push batch %s failed to prepare board: %s", batch_id, error)
        return {"status": "error", "note": batch.task_note}

    candidates_list_id = binding.get_list_id("candidates")
    if not candidates_list_id:
        batch.task_state = PubmedImportBatch.TASK_STATE_ERROR
        batch.task_note = "Candidates list is not configured for this Planka board."
        batch.save(update_fields=["task_state", "task_note", "modified"])
        return {"status": "error", "note": batch.task_note}
    created = 0
    already_pushed = 0
    failed = 0
    recreated_missing = 0

    for row in target_rows:
        if row.planka_card_id:
            try:
                existing_card = client.get_card(row.planka_card_id)
                if _is_planka_card_archived(existing_card):
                    row.planka_card_id = ""
                    row.planka_card_url = ""
                    row.planka_pushed_at = None
                    row.planka_push_error = "Planka status: previous card deleted/archived; recreating now."
                    row.save(
                        update_fields=[
                            "planka_card_id",
                            "planka_card_url",
                            "planka_pushed_at",
                            "planka_push_error",
                            "modified",
                        ]
                    )
                    recreated_missing += 1
                else:
                    existing_list_id = str(existing_card.get("listId") or "")
                    existing_list_type = list_type_map.get(existing_list_id, "")
                    if existing_list_type == "trash":
                        row.planka_card_id = ""
                        row.planka_card_url = ""
                        row.planka_pushed_at = None
                        row.planka_push_error = "Planka status: previous card deleted/archived; recreating now."
                        row.save(
                            update_fields=[
                                "planka_card_id",
                                "planka_card_url",
                                "planka_pushed_at",
                                "planka_push_error",
                                "modified",
                            ]
                        )
                        recreated_missing += 1
                    else:
                        if existing_list_id and existing_list_id != str(candidates_list_id):
                            row.planka_push_error = "Planka status: card moved from Candidates."
                            row.save(update_fields=["planka_push_error", "modified"])
                        elif row.planka_push_error:
                            row.planka_push_error = ""
                            row.save(update_fields=["planka_push_error", "modified"])
                        already_pushed += 1
                        continue
            except PlankaAPIError as error:
                if _is_planka_card_not_found_error(error):
                    row.planka_card_id = ""
                    row.planka_card_url = ""
                    row.planka_pushed_at = None
                    row.planka_push_error = "Planka status: previous card deleted/archived; recreating now."
                    row.save(
                        update_fields=[
                            "planka_card_id",
                            "planka_card_url",
                            "planka_pushed_at",
                            "planka_push_error",
                            "modified",
                        ]
                    )
                    recreated_missing += 1
                else:
                    row.planka_push_error = f"Could not verify existing Planka card: {_safe_planka_error(error)}"
                    row.save(update_fields=["planka_push_error", "modified"])
                    failed += 1
                    continue

        title, description = _build_pubmed_planka_card(row)
        try:
            card = client.create_card(candidates_list_id, title, description=description, card_type="project")
            card_id = str(card.get("id") or "").strip()
            _attach_journal_label_to_card(
                client=client,
                binding=binding,
                card_id=card_id,
                row=row,
                label_cache=label_cache,
            )
            row.planka_card_id = card_id
            base_url = (getattr(settings, "PLANKA_BASE_URL", "") or "").strip().rstrip("/")
            row.planka_card_url = f"{base_url}/cards/{card_id}" if base_url and card_id else ""
            row.planka_pushed_at = timezone.now()
            row.planka_push_error = ""
            row.save(
                update_fields=[
                    "planka_card_id",
                    "planka_card_url",
                    "planka_pushed_at",
                    "planka_push_error",
                    "modified",
                ]
            )
            created += 1
        except PlankaAPIError as error:
            row.planka_push_error = _safe_planka_error(error)
            row.save(update_fields=["planka_push_error", "modified"])
            if _is_planka_list_not_found_error(error):
                batch.task_state = PubmedImportBatch.TASK_STATE_ERROR
                batch.task_note = (
                    "Candidates list was not found in Planka. "
                    "Create or re-link a Planka board for this issue and try again."
                )
                batch.save(update_fields=["task_state", "task_note", "modified"])
                return {"status": "error", "note": batch.task_note}
            failed += 1

    batch.task_state = PubmedImportBatch.TASK_STATE_ERROR if failed else PubmedImportBatch.TASK_STATE_SUCCESS
    batch.task_note = (
        f"Push complete: {created} created, {already_pushed} already pushed, "
        f"{recreated_missing} recreated missing, {failed} failed."
    )
    batch.save(update_fields=["task_state", "task_note", "modified"])
    return {"status": "error" if failed else "success", "created": created, "failed": failed}
